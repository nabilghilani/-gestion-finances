from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header_row(ws, row_idx, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _autofit(ws, n_cols, min_width=12):
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(max_len, 40)


def export_to_excel(filepath, db):
    wb = Workbook()

    # ---- Feuille Transactions ----
    ws = wb.active
    ws.title = "Transactions"
    ws["A1"] = "Historique des transactions"
    ws["A1"].font = TITLE_FONT
    headers = ["Date", "Type", "Catégorie", "Montant", "Description"]
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 3, len(headers))

    transactions = db.get_transactions()
    for t in transactions:
        _id, date, ttype, category, amount, description = t
        ws.append([date, ttype, category, amount, description or ""])

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = THIN_BORDER
        row[3].number_format = '#,##0.00 "DA"'
    _autofit(ws, len(headers))

    # ---- Feuille Résumé mensuel ----
    ws2 = wb.create_sheet("Résumé mensuel")
    ws2["A1"] = "Résumé des 6 derniers mois"
    ws2["A1"].font = TITLE_FONT
    headers2 = ["Mois", "Revenus", "Dépenses", "Solde"]
    ws2.append([])
    ws2.append(headers2)
    _style_header_row(ws2, 3, len(headers2))

    for month, income, expense in db.get_last_n_months_totals(6):
        ws2.append([month, income, expense, income - expense])

    for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
        row[1].number_format = '#,##0.00 "DA"'
        row[2].number_format = '#,##0.00 "DA"'
        row[3].number_format = '#,##0.00 "DA"'
    _autofit(ws2, len(headers2))

    # ---- Feuille Budgets (mois courant) ----
    import datetime
    current_month = datetime.date.today().strftime("%Y-%m")
    ws3 = wb.create_sheet("Budgets")
    ws3["A1"] = f"Suivi des budgets - {current_month}"
    ws3["A1"].font = TITLE_FONT
    headers3 = ["Catégorie", "Budget prévu", "Dépensé", "Restant", "% utilisé"]
    ws3.append([])
    ws3.append(headers3)
    _style_header_row(ws3, 3, len(headers3))

    budgets = db.get_budgets(current_month)
    expenses_by_cat = dict(db.get_expenses_by_category(current_month))
    for category, budget_amount in budgets:
        spent = expenses_by_cat.get(category, 0)
        remaining = budget_amount - spent
        pct = (spent / budget_amount * 100) if budget_amount else 0
        ws3.append([category, budget_amount, spent, remaining, round(pct, 1)])

    for row in ws3.iter_rows(min_row=4, max_row=ws3.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = THIN_BORDER
        row[1].number_format = '#,##0.00 "DA"'
        row[2].number_format = '#,##0.00 "DA"'
        row[3].number_format = '#,##0.00 "DA"'
        row[4].number_format = '0.0"%"'
    _autofit(ws3, len(headers3))

    wb.save(filepath)
